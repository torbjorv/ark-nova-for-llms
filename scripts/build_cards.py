#!/usr/bin/env python3
"""Build cards.jsonl from the authoritative spreadsheet.

Reads:
  source_data/arknovaanimals_VM_v2.xlsx  — structured card data

Writes:
  cards.jsonl — one JSON object per line, every schema field present.

Run with: python scripts/build_cards.py

Note: this rebuild does not preserve hand-written `text` / `notes` from the
existing cards.jsonl. Animal text is derived from the spreadsheet ability
column where possible; sponsor / project / final-scoring text is assembled
from the spreadsheet's effect columns. Cards that exist only in the
ssimeonoff or Ender-Wiggin imports (not in the spreadsheet) won't appear
in the output — re-run those imports separately if needed.
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path
from typing import Any

import openpyxl

REPO = Path(__file__).resolve().parent.parent
XLSX = REPO / "source_data" / "arknovaanimals_VM_v2.xlsx"
OUT = REPO / "cards.jsonl"

ENCLOSURE_ICON_LETTERS = {"R": "rock_icons", "W": "water_icons"}

# Maps Type-column tokens → category tag.
# These are the 8 official animal categories from the rulebook.
CATEGORY_TOKEN_TO_TAG = {
    "predator": "predator",
    "herbivore": "herbivore",
    "bear": "bear",
    "primate": "primate",
    "reptile": "reptile",
    "bird": "bird",
    "pet": "petting-zoo",
    "sea animal": "sea-animal",
}

CATEGORY_TAGS = set(CATEGORY_TOKEN_TO_TAG.values())

CONTINENT_MAP = {
    "africa": "africa",
    "americas": "americas",
    "asia": "asia",
    "europe": "europe",
    "australia": "australia",
}

# Lookup table for the spreadsheet's requirement column. Keys are the
# spreadsheet cell value, _norm()-ed (lowercased, all whitespace including
# `\n` collapsed to single spaces, leading/trailing whitespace trimmed).
# Values are the literal `requires` tag list, with multiplicity already
# expanded (e.g. `Predator x3` → 3 × `predator`).
#
# Reputation thresholds are *not* encoded here — they're extracted separately
# by `_REP_RE` and populate the `reputation_requirement` column. Mixed strings
# like `Level II Sponsor Card, 6 Reputation` list only the non-rep tags.
#
# To support a new spreadsheet string, add one entry. Unknown strings produce
# a stderr warning during build.
REQUIREMENT_LUT: dict[str, list[str]] = {
    # --- Animal Reqs cell values ---
    "africa":                          ["africa"],
    "africa x2":                       ["africa", "africa"],
    "africa x3":                       ["africa", "africa", "africa"],
    "africa reptile":                  ["africa", "reptile"],
    "americas":                        ["americas"],
    "americas x2 animals ii":          ["americas", "americas", "animals-ii"],
    "americas x3":                     ["americas", "americas", "americas"],
    "animals ii":                      ["animals-ii"],
    "asia":                            ["asia"],
    "asia x2":                         ["asia", "asia"],
    "asia x3":                         ["asia", "asia", "asia"],
    "australia":                       ["australia"],
    "australia x2":                    ["australia", "australia"],
    "australia science":               ["australia", "science"],
    "bear animals ii":                 ["bear", "animals-ii"],
    "bird":                            ["bird"],
    "bird x2":                         ["bird", "bird"],
    "bird x3":                         ["bird", "bird", "bird"],
    "europe":                          ["europe"],
    "europe x2":                       ["europe", "europe"],
    "herbivore x2":                    ["herbivore", "herbivore"],
    "herbivore bear partner zoo":      ["herbivore", "bear", "partner-zoo"],
    "partner zoo":                     ["partner-zoo"],
    "predator x2 animals ii":          ["predator", "predator", "animals-ii"],
    "predator x3":                     ["predator", "predator", "predator"],
    "primate":                         ["primate"],
    "primate x2 animals ii":           ["primate", "primate", "animals-ii"],
    "primate x3":                      ["primate", "primate", "primate"],
    "reptile x2":                      ["reptile", "reptile"],
    "reptile x3":                      ["reptile", "reptile", "reptile"],
    "science":                         ["science"],
    "science x2":                      ["science", "science"],
    "science x2 animals ii":           ["science", "science", "animals-ii"],
    "sea animal":                      ["sea-animal"],
    "university":                      ["university"],
    "reputation 3":                    [],   # rep-only; rep_req extracted separately

    # --- Sponsor Reqs cell values ---
    "1 bird icon":                     ["bird"],
    "1 herbavore icon":                ["herbivore"],  # spreadsheet typo
    "1 predator icon":                 ["predator"],
    "1 primate icon":                  ["primate"],
    "1 reptile icon":                  ["reptile"],
    "1 research icon":                 ["science"],
    "2 partnership zoos":              ["partner-zoo", "partner-zoo"],
    "2 research icons":                ["science", "science"],
    "3 research icons":                ["science", "science", "science"],
    "4 research icons":                ["science", "science", "science", "science"],
    "2 science + 1 sea animal":        ["science", "science", "sea-animal"],
    "kiosk":                           ["kiosk"],
    "level ii sponsor card":           ["level-ii-sponsor"],
    "level ii sponsor card, 6 reputation":   ["level-ii-sponsor"],
    "level ii sponsor card; max 25 appeal":  ["level-ii-sponsor", "max-25-appeal"],
    "max. 25 appeal":                  ["max-25-appeal"],
    "sponsor ii":                      ["level-ii-sponsor"],
    "sponsor ii + 1 americas + 6 reputation": ["level-ii-sponsor", "americas"],
    "3 reputation":                    [],   # rep-only
    "6 reputation":                    [],   # rep-only
}

# Matches `Reputation N` or `N Reputation` anywhere in the cell. The count
# becomes `reputation_requirement`; the rest of the cell is looked up in
# REQUIREMENT_LUT for its tag list.
_REP_RE = re.compile(r"(?:reputation\s+(\d+)|(\d+)\s+reputation)", re.IGNORECASE)

# Named abilities printed on animal cards (value after = canonical tag).
ABILITY_NAME_MAP: dict[str, str] = {
    "sprint": "sprint",
    "pack": "pack",
    "hunter": "hunter",
    "clever": "clever",
    "inventive": "inventive",
    "full-throated": "full-throated",
    "resistance": "resistance",
    "assertion": "assertion",
    "jumping": "jumping",
    "pouch": "pouch",
    "digging": "digging",
    "flock animal": "flock",
    "flock": "flock",
    "snapping": "snapping",
    "venom": "venom",
    "sunbathing": "sunbathing",
    "hypnosis": "hypnosis",
    "constriction": "constriction",
    "scavenging": "scavenging",
    "pilfering": "pilfering",
    "posturing": "posturing",
    "perception": "perception",
    "determination": "determination",
    "peacocking": "peacocking",
    "dominance": "dominance",
    "camouflage": "camouflage",
    "symbiosis": "symbiosis",
    "helpful": "helpful",
    "trade": "trade",
    "adapt": "adapt",
    "glide": "glide",
    "scuba dive": "scuba-dive",
    "shark attack": "shark-attack",
    "cut down": "cut-down",
    "extra shift": "extra-shift",
    "mark": "mark",
    "monkey gang": "monkey-gang",
    "petting zoo animal": "petting",
}

# Parameterised abilities with sub-type (Boost: Sponsors → boost + target "sponsors").
PARAM_ABILITY_MAP = {
    "boost": "boost",
    "multiplier": "multiplier",
    "action": "action",
    "iconic animal": "iconic",
    "iconic": "iconic",
    "sponsor magnet": ("magnet", "sponsors"),
    "sea animal magnet": ("magnet", "sea-animal"),
}

SUB_TYPE_MAP = {
    "association": "association",
    "sponsors": "sponsors",
    "cards": "cards",
    "build": "build",
    "animals": "animals",
    "aminals": "animals",  # typo in source
    "africa": "africa",
    "americas": "americas",
    "asia": "asia",
    "europe": "europe",
    "australia": "australia",
    "sponsor": "sponsors",
    "card": "cards",
    "sea animal": "sea-animal",
}


def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", s.strip().lower())


# ---------------------------------------------------------------------------
# Enclosure parsing
# ---------------------------------------------------------------------------

_ENCLOSURE_ALT_RE = re.compile(r"\(([A-Z]+)\s*(\d+)\)")  # matches (RH 2), (LBA 1)
_AQ_RE = re.compile(r"Aq\s*(\d+)\s*(R)?", re.IGNORECASE)
_LRH_RE = re.compile(r"LRH\s*(\d+)", re.IGNORECASE)
_PZ_RE = re.compile(r"PZ\s*(\d+)", re.IGNORECASE)
_PAREN_NUM_RE = re.compile(r"\((\d+)\)")


ENCLOSURE_SIZE_FIELDS = (
    "standard_size",
    "reptile_house_size",
    "large_bird_aviary_size",
    "petting_zoo_size",
    "aquarium_size",
    "large_reptile_house_size",
)


def parse_enclosure(raw: Any) -> dict:
    """Return dict with keys: size, rock_icons, water_icons, and the six per-enclosure size fields.

    `size` is the scoring/category size: equal to `standard_size` for land animals
    and to the parenthesised number for sea animals. Each `*_size` field is set
    only when the card has that enclosure kind printed; others stay None.
    `rock_icons` / `water_icons` are integer counts of those enclosure-requirement
    icons printed on the card.
    """
    out: dict = {
        "size": None,
        "rock_icons": 0,
        "water_icons": 0,
    }
    for f in ENCLOSURE_SIZE_FIELDS:
        out[f] = None
    if raw is None:
        return out
    s = str(raw).strip()
    if not s:
        return out

    # Petting zoo: "PZ 1"
    if s.upper().startswith("PZ"):
        m = _PZ_RE.search(s)
        if m:
            n = int(m.group(1))
            out["petting_zoo_size"] = n
            out["size"] = n
        return out

    # Sea animal: "(N)  Aq M [R]" with optional "/ LRH P" suffix
    if s.startswith("("):
        pm = _PAREN_NUM_RE.search(s)
        aqm = _AQ_RE.search(s)
        lrhm = _LRH_RE.search(s)
        if pm:
            out["size"] = int(pm.group(1))
        if aqm:
            out["aquarium_size"] = int(aqm.group(1))
            if aqm.group(2):  # trailing 'R'
                out["rock_icons"] += 1
        if lrhm:
            out["large_reptile_house_size"] = int(lrhm.group(1))
        return out

    # Penguin-style dual enclosure: "3RW / Aq 2"
    if "/" in s and "Aq" in s:
        left, right = [p.strip() for p in s.split("/", 1)]
        main = parse_enclosure(left)
        altm = _AQ_RE.search(right)
        for f in ENCLOSURE_SIZE_FIELDS:
            if main[f] is not None:
                out[f] = main[f]
        out["size"] = main["size"]
        if altm:
            out["aquarium_size"] = int(altm.group(1))
        out["rock_icons"] = main["rock_icons"]
        out["water_icons"] = main["water_icons"]
        return out

    # Land animal with optional (RH N) or (LBA N) alt: e.g. "5W (RH 3)", "4 (LBA 1)", "3R"
    alt_match = _ENCLOSURE_ALT_RE.search(s)
    if alt_match:
        code = alt_match.group(1)
        alt_n = int(alt_match.group(2))
        if code == "RH":
            out["reptile_house_size"] = alt_n
        elif code == "LBA":
            out["large_bird_aviary_size"] = alt_n
        head = s[: alt_match.start()].strip()
    else:
        head = s

    # head is now like "5", "3R", "2RR", "5WR", "4WW"
    m = re.match(r"(\d+)\s*([RW]*)", head)
    if m:
        n = int(m.group(1))
        out["standard_size"] = n
        out["size"] = n
        for ch in m.group(2):
            if ch in ENCLOSURE_ICON_LETTERS:
                out[ENCLOSURE_ICON_LETTERS[ch]] += 1
    return out


# ---------------------------------------------------------------------------
# Type / continent columns
# ---------------------------------------------------------------------------

def parse_type_column(raw: Any) -> list[str]:
    """Return the list of category tags printed on the card.

    Handles 'Predator', 'Predator/Bear', 'Predator x2', 'Sea Animal / Reptile', 'Sea Animal 2'.
    Duplicate tags encode multiplicity: 'Predator x2' → ['predator','predator'].
    """
    cats: list[str] = []
    if raw is None:
        return cats
    s = str(raw).strip().lower()
    # Split on '/' for multi-category
    parts = [p.strip() for p in s.split("/")]
    for p in parts:
        # Extract multiplier (xN or trailing " 2")
        mult = 1
        m = re.search(r"x\s*(\d+)\s*$", p)
        if m:
            mult = int(m.group(1))
            p = p[: m.start()].strip()
        else:
            m = re.search(r"\s(\d+)\s*$", p)
            if m:
                mult = int(m.group(1))
                p = p[: m.start()].strip()
        p = p.strip()
        tag = CATEGORY_TOKEN_TO_TAG.get(p)
        if tag:
            cats.extend([tag] * mult)
    return cats


def parse_continent_column(raw: Any) -> list[str]:
    if raw is None:
        return []
    s = str(raw).strip().lower()
    if s == "none":
        return []
    out: list[str] = []
    for part in re.split(r"[\n,/]", s):
        part = part.strip()
        if not part:
            continue
        mult = 1
        m = re.search(r"x\s*(\d+)\s*$", part)
        if m:
            mult = int(m.group(1))
            part = part[: m.start()].strip()
        tag = CONTINENT_MAP.get(part)
        if tag:
            out.extend([tag] * mult)
    return out


# ---------------------------------------------------------------------------
# Reqs column
# ---------------------------------------------------------------------------

def parse_reqs_column(raw: Any) -> tuple[list[str], int | None]:
    """Return (requires tags, reputation_requirement).

    The cell value is normalized (lowercased, whitespace collapsed) and looked
    up directly in REQUIREMENT_LUT. Reputation thresholds are extracted by
    `_REP_RE` independently of the LUT.
    """
    if raw is None:
        return [], None
    key = _norm(str(raw))
    if not key:
        return [], None

    rep_match = _REP_RE.search(key)
    rep_req = int(rep_match.group(1) or rep_match.group(2)) if rep_match else None

    tags = REQUIREMENT_LUT.get(key)
    if tags is None:
        print(f"WARN: unknown requirements string: {raw!r}", file=sys.stderr)
        return [], rep_req

    return list(tags), rep_req


# ---------------------------------------------------------------------------
# Ability column
# ---------------------------------------------------------------------------

def parse_ability_column(raw: Any) -> tuple[list[str], dict[str, int], dict[str, str]]:
    """Return (tags, ability_levels, ability_targets)."""
    tags: list[str] = []
    levels: dict[str, int] = {}
    targets: dict[str, str] = {}

    if raw is None:
        return tags, levels, targets
    s = str(raw).strip()
    if not s or s.lower() == "none":
        return tags, levels, targets

    # Split on "\n", "/", ","
    parts = [p.strip() for p in re.split(r"[\n,/]+", s) if p.strip()]

    for part in parts:
        tag, lvl, tgt = _parse_ability_part(part)
        if tag is None:
            continue
        tags.append(tag)
        if lvl is not None:
            levels[tag] = lvl
        if tgt is not None:
            targets[tag] = tgt

    return tags, levels, targets


def _parse_ability_part(raw: str) -> tuple[str | None, int | None, str | None]:
    s = raw.strip()
    low = _norm(s).replace(":", " ")
    # Form: "Boost: Association", "Multiplier: Sponsors", "Iconic Animal: Europe"
    for key in PARAM_ABILITY_MAP:
        if low.startswith(key + " ") or low == key:
            rest = low[len(key):].strip()
            mapping = PARAM_ABILITY_MAP[key]
            if isinstance(mapping, tuple):
                tag, tgt = mapping
                return tag, None, tgt
            tgt_tag = SUB_TYPE_MAP.get(rest)
            if tgt_tag:
                return mapping, None, tgt_tag
            return mapping, None, None

    # Form: "Sprint 3", "Hunter 4", "Posturing 1"
    m = re.match(r"^([a-z \-]+?)\s+(\d+|x)\s*$", low)
    if m:
        name = m.group(1).strip()
        lvl_raw = m.group(2)
        tag = ABILITY_NAME_MAP.get(name)
        if tag:
            lvl = None if lvl_raw == "x" else int(lvl_raw)
            return tag, lvl, None

    # Bare name
    tag = ABILITY_NAME_MAP.get(low)
    if tag:
        return tag, None, None

    return None, None, None


# ---------------------------------------------------------------------------
# Bonuses (A/C/R)
# ---------------------------------------------------------------------------

def parse_bonuses(raw: Any) -> tuple[int | None, int | None, int | None]:
    if raw is None:
        return None, None, None
    s = str(raw).strip()
    m = re.match(r"(\d+)\s*/\s*(\d+)\s*/\s*(\d+)", s)
    if not m:
        return None, None, None
    a, c, r = int(m.group(1)), int(m.group(2)), int(m.group(3))
    return (a if a else None), (c if c else None), (r if r else None)


# ---------------------------------------------------------------------------
# Card builders
# ---------------------------------------------------------------------------

EMPTY = {
    "rock_icons": 0, "water_icons": 0, "continents": [], "categories": [], "size": None,
    "abilities": [], "requires": [], "provides": [], "triggers": [],
    "appeal": None, "conservation_points": None, "strength": None,
    "reputation_requirement": None, "reputation_reward": None, "money_cost": None,
    "text": "", "notes": None,
    "standard_size": None, "reptile_house_size": None, "large_bird_aviary_size": None,
    "petting_zoo_size": None, "aquarium_size": None, "large_reptile_house_size": None,
    "reef_ability": None, "wave_icon": False,
    "ability_levels": {}, "ability_targets": {},
    "tier_thresholds": [], "tier_rewards": [],
}


def new_card(**kw) -> dict:
    card = {**EMPTY, **kw}
    return card


def _titlecase(name: str) -> str:
    # "GREVY'S ZEBRA" → "Grevy's Zebra"; handles parentheses and hyphens.
    if not name:
        return name
    out = []
    for word in re.split(r"(\s+|[-/])", name):
        if word.strip() and word not in "-/":
            out.append(word[:1].upper() + word[1:].lower())
        else:
            out.append(word)
    return "".join(out)


# ---------------------------------------------------------------------------
# Sheet readers
# ---------------------------------------------------------------------------

def read_animals(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    cards = []
    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        num = int(row[0])
        name_raw = row[1]
        if not name_raw:
            continue
        latin = row[2]
        enc_raw = row[3]
        cost = row[4]
        type_raw = row[5]
        cont_raw = row[6]
        reqs_raw = row[7]
        abil_raw = row[8]
        bonus_raw = row[9]
        reef = row[10]
        wave = row[11]
        mw_flag = row[12]

        is_mw = (mw_flag == "MW")
        set_name = "marine-worlds" if is_mw else "base"
        prefix = "MW" if is_mw else "AN"
        cid = f"{prefix}-{num:03d}"

        enc = parse_enclosure(enc_raw)
        categories = parse_type_column(type_raw)
        continents = parse_continent_column(cont_raw)
        req_tags, rep_req = parse_reqs_column(reqs_raw)
        abil_tags, abil_levels, abil_targets = parse_ability_column(abil_raw)
        appeal, cp, rep_reward = parse_bonuses(bonus_raw)

        abilities = abil_tags
        wave_bool = bool(wave)

        notes = f"Latin: {latin}" if latin else None
        text = ""

        card = new_card(
            id=cid,
            name=_titlecase(name_raw),
            set=set_name,
            type="animal",
            rock_icons=enc["rock_icons"],
            water_icons=enc["water_icons"],
            continents=continents,
            categories=categories,
            size=enc["size"],
            abilities=abilities,
            requires=req_tags,
            provides=[],
            triggers=[],
            appeal=appeal,
            conservation_points=cp,
            reputation_requirement=rep_req,
            reputation_reward=rep_reward,
            money_cost=cost if isinstance(cost, int) else None,
            text=text,
            notes=notes,
            standard_size=enc["standard_size"],
            reptile_house_size=enc["reptile_house_size"],
            large_bird_aviary_size=enc["large_bird_aviary_size"],
            petting_zoo_size=enc["petting_zoo_size"],
            aquarium_size=enc["aquarium_size"],
            large_reptile_house_size=enc["large_reptile_house_size"],
            reef_ability=str(reef).strip() if reef else None,
            wave_icon=wave_bool,
            ability_levels=abil_levels,
            ability_targets=abil_targets,
        )
        cards.append(card)
    return cards


def read_sponsors(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    cards = []
    for row in rows[1:]:
        if not row or row[0] is None or not isinstance(row[0], int):
            continue
        num = row[0]
        name = row[1]
        strength = row[2]
        reqs_raw = row[3]
        icons_gained_raw = row[4]
        instant = row[5]
        ongoing = row[6]
        endgame = row[7]
        wave = row[8]
        mw_raw = row[9]

        is_mw = (mw_raw == "MW")
        # Sponsors: if marked MW, it's a replacement → MW-###. Otherwise base → AN-###.
        prefix = "MW" if is_mw else "AN"
        set_name = "marine-worlds" if is_mw else "base"
        cid = f"{prefix}-{num:03d}"

        # Parse requirements
        req_tags, rep_req = parse_reqs_column(reqs_raw)

        # Parse provided icons
        provides = parse_provides(icons_gained_raw)

        triggers = []
        if instant:
            triggers.append("immediate")
        if ongoing:
            triggers.append("ongoing")
        if endgame:
            triggers.append("end")

        # Text assembled from instant/ongoing/endgame segments
        parts = [str(x).strip() for x in (instant, ongoing, endgame) if x]
        text = " / ".join(parts) if parts else ""
        notes = None

        abilities: list[str] = []
        # Tag science/research sponsors with `science` ability when they produce Research icons.
        if provides and any(p == "science" for p in provides):
            abilities.append("science")

        card = new_card(
            id=cid,
            name=name,
            set=set_name,
            type="sponsor",
            strength=strength if isinstance(strength, int) else None,
            requires=req_tags,
            provides=provides,
            triggers=triggers,
            reputation_requirement=rep_req,
            text=text,
            notes=notes,
            abilities=abilities,
            wave_icon=bool(wave),
        )
        cards.append(card)
    return cards


def parse_provides(raw: Any) -> list[str]:
    """Parse 'Icons Gained' column into a list of tag, duplicated for multiplicity."""
    if raw is None:
        return []
    s = str(raw)
    provides: list[str] = []
    # Strip parenthesised alternates
    s = re.sub(r"\([^)]*\)", "", s)
    for part in re.split(r"[+,\n]+", s):
        part = part.strip()
        if not part:
            continue
        m = re.match(r"(\d+)\s+(.+)", part)
        if not m:
            continue
        count = int(m.group(1))
        name = _norm(m.group(2))
        # normalise 'Research' → science
        if name in ("research", "science", "sceince"):
            tag = "science"
        elif name in CATEGORY_TOKEN_TO_TAG:
            tag = CATEGORY_TOKEN_TO_TAG[name]
        elif name in CONTINENT_MAP:
            tag = CONTINENT_MAP[name]
        elif name == "rock":
            tag = "rock"
        elif name == "water":
            tag = "water"
        elif name in ("rocks", "waters"):
            tag = name[:-1]
        elif name == "petting zoo animal":
            tag = "petting-zoo"
        elif name == "x-token":
            tag = None
        else:
            tag = None
        if tag:
            provides.extend([tag] * count)
    return provides


def read_conservation(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    cards = []
    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        num = int(row[0])
        name = row[1]
        deck = row[2]  # Base / Zoo
        activity = row[3]
        size_req_raw = row[4]
        cp_raw = row[5]
        rep = row[6]
        requirements_text = row[7]
        mw_flag = row[8]

        tier_thr = parse_tier(size_req_raw)
        tier_rew = parse_tier(cp_raw)

        # Map activity → requires tag
        activity_tag = {
            "Collection": "collection-activity",
            "Release": "release-activity",
            "Partnership": "partnership-activity",
        }.get(activity)

        requires = [activity_tag] if activity_tag else []
        triggers = []
        if activity == "Release":
            triggers.append("on-release")
        elif activity == "Partnership":
            triggers.append("on-partnership")

        # Continent / animal-category requirements from the free-text field
        extra_tags = _infer_project_requires(requirements_text)
        requires.extend(extra_tags)
        categories = _project_categories(requires)

        rep_reward = None
        if isinstance(rep, int):
            rep_reward = rep
        # '2/2/0' tiered reputation — we record first tier as reward
        elif isinstance(rep, str):
            m = re.match(r"(\d+)", rep)
            if m:
                rep_reward = int(m.group(1))

        is_mw = (mw_flag == "MW")
        # Base-game printed card → AN-###. If the Deck is 'Base' and MW flag is present,
        # that means this row is the MW REPLACEMENT for the base card → MW-### (and we
        # also keep the base version via backup-only if it existed).
        # Simpler convention: if deck==Zoo, always AN-### (zoo-pack version); if deck==Base
        # and MW flag, MW-###; else AN-###.
        if deck == "Zoo":
            prefix = "AN"
            set_name = "base"  # zoo-pack projects shipped with base? Actually they're from the zoo map pack
        elif is_mw:
            prefix = "MW"
            set_name = "marine-worlds"
        else:
            prefix = "AN"
            set_name = "base"
        cid = f"{prefix}-{num:03d}"

        text = requirements_text or ""
        notes = None

        card = new_card(
            id=cid,
            name=name.replace("*", "").strip() if name else name,
            set=set_name,
            type="conservation-project",
            categories=categories,
            requires=requires,
            triggers=triggers,
            reputation_reward=rep_reward,
            tier_thresholds=tier_thr,
            tier_rewards=tier_rew,
            text=text,
            notes=notes,
        )
        cards.append(card)
    return cards


def _infer_project_requires(text: str | None) -> list[str]:
    if not text:
        return []
    low = text.lower()
    tags: list[str] = []
    for kw, tag in (
        ("africa", "africa"),
        ("americas", "americas"),
        ("asia", "asia"),
        ("europe", "europe"),
        ("australia", "australia"),
        ("bird", "bird"),
        ("predator", "predator"),
        ("reptile", "reptile"),
        ("primate", "primate"),
        ("herbavore", "herbivore"),
        ("herbivore", "herbivore"),
        ("water icon", "water"),
        ("rock icon", "rock"),
        ("research icon", "science"),
    ):
        if kw in low and tag not in tags:
            tags.append(tag)
    return tags


def _project_categories(requires_tags: list[str]) -> list[str]:
    """The project card prints its category icon. If `requires` includes a
    category tag (because the project demands those icons in the zoo), the
    same category appears on the project card itself."""
    return [t for t in requires_tags if t in CATEGORY_TAGS]


def parse_tier(raw: Any) -> list[int]:
    """Parse tier strings like '5/4/2' or '1/2/3/4'.

    Reject range-style strings ('1 to 4', '1  to 4') — those are not tiered schedules.
    """
    if raw is None:
        return []
    s = str(raw).strip()
    # Drop parenthesised alternates (e.g. "(1/3/5/6)*")
    s = re.sub(r"\([^)]*\)", "", s).strip()
    if re.search(r"\bto\b", s, flags=re.IGNORECASE):
        return []
    m = re.findall(r"\d+", s)
    return [int(x) for x in m] if m else []


def read_final_scoring(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    cards = []
    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        num = int(row[0]) if isinstance(row[0], int) else int(row[0])
        name = row[1]
        req_raw = row[2]
        cp_raw = row[3]
        text_raw = row[4]
        extra = row[5]
        mw_flag = row[6]

        tier_thr = parse_tier(req_raw)
        tier_rew = parse_tier(cp_raw)

        is_mw = (mw_flag == "MW")
        prefix = "MW" if is_mw else "AN"
        set_name = "marine-worlds" if is_mw else "base"
        cid = f"{prefix}-{num:03d}"

        text = text_raw or ""
        notes = extra if extra else None

        card = new_card(
            id=cid,
            name=(name or "").replace("*", "").strip(),
            set=set_name,
            type="final-scoring",
            tier_thresholds=tier_thr,
            tier_rewards=tier_rew,
            text=text,
            notes=notes,
        )
        cards.append(card)
    return cards


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> int:
    if not XLSX.exists():
        print(f"ERROR: {XLSX} not found", file=sys.stderr)
        return 2

    wb = openpyxl.load_workbook(XLSX, data_only=True)

    all_cards: list[dict] = []
    all_cards += read_final_scoring(wb["Final Scoring"])
    all_cards += read_conservation(wb["Conservation"])
    all_cards += read_sponsors(wb["Sponsors"])
    all_cards += read_animals(wb["Animals"])

    all_cards.sort(key=lambda c: (c["type"], c["id"]))

    with OUT.open("w") as f:
        for card in all_cards:
            f.write(json.dumps(card, ensure_ascii=False) + "\n")

    print(f"Wrote {len(all_cards)} cards to {OUT}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
